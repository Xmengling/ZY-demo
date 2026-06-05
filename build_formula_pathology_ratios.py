from __future__ import annotations

import json
import re
import sqlite3
from collections import defaultdict
from pathlib import Path

from docx import Document
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


BASE = Path(r"D:\AI\demo")
TCM_JSON = BASE / "ai-medical-consultant" / "backend" / "data" / "tcm_knowledge.json"
HERB_DOCX = BASE / "tcm_rag_demo" / "data" / "李冠杰经方病理辨证体系经方药物集萃（目录)(1).docx"
JINGFANG_DB = BASE / "ai-medical-consultant" / "backend" / "data" / "jingfang.sqlite3"
OUT_XLSX = BASE / "经方方剂组成与病理比例整理表.xlsx"

SOURCE_KEY = "经方病理辨证体系"

CANON_LABELS = {
    "表实",
    "表虚",
    "里实",
    "里虚",
    "里寒",
    "里热",
    "半热",
    "水实",
    "水虚",
    "血实",
    "血虚",
    "气实",
    "气虚",
    "阴证",
}

LABEL_ALIASES = {
    "表证": "表证",
    "阴性": "阴证",
    "阴证": "阴证",
    "辅助阴性": "阴证",
    "血瘀": "血实",
    "瘀血": "血实",
    "津虚": "水虚",
    "水虚": "水虚",
}

HERB_ALIASES = {
    "炙甘草": "甘草",
    "甘草炙": "甘草",
    "生甘草": "甘草",
    "生姜切": "生姜",
    "干姜": "干姜",
    "姜": "生姜",
    "大枣擘": "大枣",
    "大枣": "大枣",
    "半夏洗": "半夏",
    "葶苈": "葶苈子",
    "葶苈子": "葶苈子",
    "芍药": "芍药",
    "白芍": "芍药",
    "赤芍": "芍药",
    "茵陈": "茵陈蒿",
    "茵陈蒿": "茵陈蒿",
    "地黄": "生地",
    "生地黄": "生地",
    "生地": "生地",
    "熟地黄": "生地",
    "桂枝": "桂枝",
    "桂": "桂枝",
    "牡蛎熬": "牡蛎",
    "栝蒌根": "天花粉",
    "瓜蒌根": "天花粉",
    "括楼根": "天花粉",
    "黄耆": "黄芪",
    "薏苡": "薏苡仁",
    "苡仁": "薏苡仁",
}


MANUAL_ROLES: dict[str, dict[str, list[str]]] = {
    "四逆散": {
        "柴胡": ["半热"],
        "枳实": ["里实", "气实"],
        "芍药": ["里实", "血虚"],
        "甘草": [],
    },
    "小柴胡汤": {
        "柴胡": ["半热"],
        "黄芩": ["半热"],
        "人参": ["里虚"],
        "甘草": ["里虚"],
        "大枣": ["里虚"],
        "半夏": ["水实"],
        "生姜": ["水实"],
    },
    "半夏厚朴汤": {
        "半夏": ["水实"],
        "茯苓": ["水实"],
        "生姜": ["水实"],
        "厚朴": ["气实"],
        "苏叶": ["气实"],
    },
    "麻杏石甘汤": {
        "麻黄": ["表实"],
        "杏仁": ["表实"],
        "石膏": ["里热"],
        "甘草": [],
    },
    "桂枝汤": {
        "桂枝": ["表虚"],
        "生姜": ["表虚"],
        "甘草": ["里虚"],
        "大枣": ["里虚"],
        "芍药": ["血虚"],
    },
    "枳实芍药散": {
        "枳实": ["气实", "里实"],
        "芍药": ["血虚", "里实"],
    },
}


def normalize_label(label: str) -> str:
    label = label.strip().strip("。；;，,、")
    return LABEL_ALIASES.get(label, label)


def normalize_herb(name: str) -> str:
    name = re.sub(r"[（）()一二三四五六七八九十半升两枚分各切洗炙熬去皮尖]+$", "", name.strip())
    name = name.strip(" ：:，,、。；;")
    return HERB_ALIASES.get(name, name)


def parse_herb_labels() -> dict[str, list[str]]:
    doc = Document(str(HERB_DOCX))
    labels: dict[str, list[str]] = {}
    for para in doc.paragraphs:
        text = para.text.strip()
        if SOURCE_KEY not in text:
            continue
        raw = text.split("：", 1)[1] if "：" in text else text.split(":", 1)[-1]
        raw = raw.strip().rstrip("。.")
        if "：" in raw:
            herb, rest = raw.split("：", 1)
        elif "，" in raw:
            herb, rest = raw.split("，", 1)
        elif "," in raw:
            herb, rest = raw.split(",", 1)
        else:
            continue
        herb = normalize_herb(herb)
        parts = [normalize_label(x) for x in re.split(r"[\s、，,]+", rest) if x.strip()]
        dedup = []
        for part in parts:
            if part and part not in dedup:
                dedup.append(part)
        labels[herb] = dedup
    return labels


def parse_pathologies(content: str) -> list[str]:
    m = re.search(r"【病理】(.+?)(?:。|\n|【)", content, re.S)
    if not m:
        return []
    raw = m.group(1).strip()
    parts = [normalize_label(x) for x in re.split(r"[\s、，,；;。]+", raw) if x.strip()]
    dedup = []
    for p in parts:
        if p and p not in dedup:
            dedup.append(p)
    return dedup


def parse_composition_line(line: str) -> list[tuple[str, float]]:
    line = line.replace("。", " ").replace("，", " ").replace(",", " ")
    line = re.sub(r"（.*?）|\(.*?\)", " ", line)
    pairs = []
    for herb, dose in re.findall(r"([\u4e00-\u9fff]+)\s*(\d+(?:\.\d+)?)", line):
        herb = normalize_herb(herb)
        if not herb or len(herb) > 8:
            continue
        pairs.append((herb, float(dose)))
    # collapse duplicates
    totals: dict[str, float] = defaultdict(float)
    for herb, dose in pairs:
        totals[herb] += dose
    return list(totals.items())


def extract_composition(content: str) -> tuple[list[tuple[str, float]], str]:
    candidates: list[str] = []
    m = re.search(r"【方解】([^\n。]+)", content)
    if m:
        candidates.append(m.group(1))
    m = re.search(r"散剂推荐比例：\s*\n?([^\n]+)", content)
    if m:
        candidates.append(m.group(1))
    m = re.search(r"汤剂推荐比例：\s*\n?([^\n]+)", content)
    if m and not candidates:
        candidates.append(m.group(1))
    for line in candidates:
        comp = parse_composition_line(line)
        if len(comp) >= 2:
            return comp, line.strip()
    return [], ""


def pathology_intersection(
    formula_paths: list[str], herb_labels: list[str], herb: str
) -> list[str]:
    fset = set(formula_paths)
    hset = set(herb_labels)
    roles = []
    # Direct matches.
    for label in formula_paths:
        if label in hset and label in CANON_LABELS:
            roles.append(label)
    # Source often says "表证"; decide by formula's concrete表实/表虚.
    if "表证" in hset:
        for label in ("表实", "表虚"):
            if label in fset and label not in roles:
                roles.append(label)
    # If formula says 阴性/阴证 and herb has 附子类阴证 label.
    if "阴证" in fset and "阴证" in hset and "阴证" not in roles:
        roles.append("阴证")
    # Keep only target labels.
    roles = [r for r in roles if r in CANON_LABELS]
    return roles


def format_num(n: float) -> str:
    return str(int(n)) if abs(n - int(n)) < 1e-9 else f"{n:g}"


def build_rows():
    herb_labels = parse_herb_labels()
    data = json.loads(TCM_JSON.read_text(encoding="utf-8"))
    formula_rows = []
    detail_rows = []
    needs_review = []

    for item in data:
        title = item.get("title", "").strip()
        content = item.get("content", "")
        if not title or item.get("category") != "经方方证":
            continue
        formula_paths = parse_pathologies(content)
        comp, comp_source = extract_composition(content)
        if not comp:
            needs_review.append([title, "未能提取组成", formula_paths, comp_source])
            continue

        role_points: dict[str, float] = defaultdict(float)
        role_drugs: dict[str, list[str]] = defaultdict(list)
        unassigned = []
        confidence = "自动"
        if title in MANUAL_ROLES:
            confidence = "手工校正"

        for herb, dose in comp:
            base = normalize_herb(herb)
            if title in MANUAL_ROLES and base in MANUAL_ROLES[title]:
                roles = MANUAL_ROLES[title][base]
            else:
                roles = pathology_intersection(formula_paths, herb_labels.get(base, []), base)
            if not roles:
                unassigned.append(f"{herb}{format_num(dose)}")
            for role in roles:
                role_points[role] += dose
                role_drugs[role].append(f"{herb}{format_num(dose)}")
                detail_rows.append(
                    [
                        title,
                        herb,
                        dose,
                        role,
                        "手工校正" if title in MANUAL_ROLES and base in MANUAL_ROLES[title] else "药物标签∩方剂病理",
                        " ".join(herb_labels.get(base, [])),
                        " ".join(formula_paths),
                    ]
                )

        total = sum(d for _, d in comp)
        role_total = sum(role_points.values())
        if role_total:
            ratio_parts = [
                f"{k} {role_points[k] / role_total * 100:.1f}%"
                for k in sorted(role_points, key=lambda x: (-role_points[x], x))
            ]
            mapping_parts = [
                f"{k}{format_num(role_points[k])}：{' '.join(role_drugs[k])}"
                for k in sorted(role_points, key=lambda x: (-role_points[x], x))
            ]
        else:
            ratio_parts = []
            mapping_parts = []
            needs_review.append([title, "未能推断病理对应药物", formula_paths, comp_source])

        formula_rows.append(
            [
                title,
                "+".join(f"{h}{format_num(d)}" for h, d in comp),
                total,
                " ".join(formula_paths),
                role_total,
                "、".join(ratio_parts),
                "\n".join(mapping_parts),
                "、".join(unassigned),
                confidence,
                comp_source,
            ]
        )

    return formula_rows, detail_rows, herb_labels, needs_review


def add_db_rows(formula_rows):
    """Append formulas from jingfang.sqlite3 when not already in tcm_knowledge."""
    existing = {r[0] for r in formula_rows}
    if not JINGFANG_DB.exists():
        return
    con = sqlite3.connect(str(JINGFANG_DB))
    try:
        for fid, payload in con.execute("select id, payload from formulas"):
            try:
                data = json.loads(payload)
            except Exception:
                continue
            name = data.get("name") or data.get("title") or fid
            if name in existing:
                continue
            comp_text = data.get("composition") or data.get("ingredients") or ""
            path_text = data.get("pathology") or data.get("bingli") or ""
            if not isinstance(comp_text, str):
                comp_text = json.dumps(comp_text, ensure_ascii=False)
            if not isinstance(path_text, str):
                path_text = json.dumps(path_text, ensure_ascii=False)
            formula_rows.append(
                [
                    name,
                    comp_text,
                    "",
                    path_text,
                    "",
                    "",
                    "",
                    "",
                    "数据库已有记录，需人工拆分比例",
                    "jingfang.sqlite3",
                ]
            )
    finally:
        con.close()


def dedupe_formula_rows(formula_rows):
    """Keep one best summary row for each formula name."""
    best = {}

    def score(row):
        value = 0
        if row[3]:
            value += 100
        if row[5]:
            value += 50
        if row[6]:
            value += 50
        try:
            value += float(row[4] or 0)
        except Exception:
            pass
        return value

    for row in formula_rows:
        name = row[0]
        if name not in best or score(row) > score(best[name]):
            best[name] = row
    return list(best.values())


def dedupe_detail_rows(detail_rows):
    """Keep one detail row for each formula-herb-dose-pathology role."""
    best = {}

    def score(row):
        return int(bool(row[5])) + int(bool(row[6]))

    for row in detail_rows:
        key = tuple(row[:4])
        if key not in best or score(row) > score(best[key]):
            best[key] = row
    return list(best.values())


def write_workbook(formula_rows, detail_rows, herb_labels, needs_review):
    wb = Workbook()
    ws = wb.active
    ws.title = "方剂病理比例总表"
    headers = [
        "方剂",
        "药物组成",
        "单方总量",
        "原资料病理",
        "病理点数合计",
        "病理比例",
        "病理对应药物",
        "未分配/调和药物",
        "整理方式",
        "组成来源原文",
    ]
    ws.append(headers)
    for row in formula_rows:
        ws.append(row)

    ws2 = wb.create_sheet("药物角色明细")
    ws2.append(["方剂", "药物", "剂量", "承担病理", "依据", "药物原始标签", "方剂原始病理"])
    for row in detail_rows:
        ws2.append(row)

    ws3 = wb.create_sheet("药物病理来源")
    ws3.append(["药物", "原始病理标签"])
    for herb, labels in sorted(herb_labels.items()):
        ws3.append([herb, " ".join(labels)])

    ws4 = wb.create_sheet("待校正记录")
    ws4.append(["方剂", "问题", "原资料病理", "组成来源原文"])
    for title, issue, paths, source in needs_review:
        ws4.append([title, issue, " ".join(paths), source])

    info = wb.create_sheet("说明")
    info.append(["项目", "说明"])
    info.append(["比例算法", "病理比例 = 该病理对应药物剂量合计 / 所有已分配病理点数合计。一个药物承担多个病理时会重复计入病理点数。"])
    info.append(["手工校正", "四逆散、小柴胡汤、半夏厚朴汤、麻杏石甘汤、桂枝汤、枳实芍药散等已按方解角色校正。"])
    info.append(["自动推断", "其他方剂用“药物集萃病理标签 ∩ 方剂【病理】标签”推断，建议后续逐方人工校正。"])
    info.append(["未分配", "甘草、大枣等调和/辅助药，或资料中病理角色不清者，放入未分配/调和药物。"])

    for sheet in wb.worksheets:
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="0B2545")
            cell.fill = PatternFill("solid", fgColor="E8EEF5")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        widths = {
            "方剂病理比例总表": [18, 42, 10, 28, 12, 42, 58, 28, 18, 52],
            "药物角色明细": [18, 14, 10, 14, 18, 32, 28],
            "药物病理来源": [18, 42],
            "待校正记录": [20, 30, 30, 60],
            "说明": [18, 80],
        }.get(sheet.title, [])
        for i, width in enumerate(widths, start=1):
            sheet.column_dimensions[get_column_letter(i)].width = width
        sheet.freeze_panes = "A2"
    wb.save(OUT_XLSX)


def main():
    formula_rows, detail_rows, herb_labels, needs_review = build_rows()
    formula_rows = dedupe_formula_rows(formula_rows)
    detail_rows = dedupe_detail_rows(detail_rows)
    add_db_rows(formula_rows)
    write_workbook(formula_rows, detail_rows, herb_labels, needs_review)
    print("formula_rows", len(formula_rows))
    print("detail_rows", len(detail_rows))
    print("herb_labels", len(herb_labels))
    print("needs_review", len(needs_review))
    print(OUT_XLSX)


if __name__ == "__main__":
    main()
